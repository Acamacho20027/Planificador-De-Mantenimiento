"""
Microservicio FastAPI para procesar bitácoras y generar un único Excel fijo.
Contexto actual: un solo cliente/sede y un único archivo que se actualiza diariamente.
"""
import glob
import json
import logging
import os
import re
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from fastapi import Body, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

BASE_DIR = Path(__file__).resolve().parent
BITACORA_SOURCE_DIR = BASE_DIR / "bitacoras"
BITACORA_OUTPUT_DIR = BASE_DIR / "output"
BITACORA_DEFAULT_CLIENTE = os.getenv("BITACORA_DEFAULT_CLIENTE", "Cliente Principal")
CLIENTS_FILE = BITACORA_OUTPUT_DIR / "clientes.json"

BITACORA_SOURCE_DIR.mkdir(parents=True, exist_ok=True)
BITACORA_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
logger = logging.getLogger(__name__)

EXPECTED_FIELDS = []

FIELD_ALIASES = {
    "fecha": "fecha",
    "fecha_ingreso": "fecha",
    "f": "fecha",
    "cliente": "cliente",
    "empresa": "cliente",
    "compania": "cliente",
    "tecnico": "tecnico",
    "tecnicos": "tecnico",
    "responsable": "tecnico",
    "tarea": "tarea",
    "actividad": "tarea",
    "trabajo": "tarea",
    "estado": "estado",
    "estatus": "estado",
    "observaciones": "observaciones",
    "obs": "observaciones",
    "comentarios": "observaciones",
}

DATE_PATTERNS = [
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%Y/%m/%d",
    "%Y%m%d",
]

INVISIBLE_CHARS_RE = re.compile(r"[\u200b-\u200f\u202a-\u202e\u2060-\u2064\ufeff]")

WHATSAPP_LINE_REGEX = re.compile(
    r"""
    ^\s*
    \[?                                     # bracket opcional
    (?P<fecha>\d{1,2}/\d{1,2}/\d{2,4})     # fecha en formato d/m/yy(yy)
    \s*,\s*
    (?P<hora>\d{1,2}:\d{2}(?::\d{2})?)     # hora con opcional segundos
    \s*
    (?P<ampm>AM|PM)?                         # AM/PM opcional
    \]?\s*
    (?:~\s*)?                               # prefijo ~ opcional
    (?:-\s*)?                               # separador guion opcional
    (?P<autor>[^:]+?)                        # autor hasta el siguiente ':'
    \s*:\s*
    (?P<mensaje>.+)                          # mensaje
    $
    """,
    re.IGNORECASE | re.VERBOSE,
)

DATE_FORMATS = ["%d/%m/%y", "%d/%m/%Y"]
TIME_FORMATS_AMPM = ["%I:%M:%S %p", "%I:%M %p"]
TIME_FORMATS_24H = ["%H:%M:%S", "%H:%M"]

NOISE_KEYWORDS = {
    "imagen omitida",
    "audio omitido",
    "video omitido",
    "mensaje eliminado",
    "se eliminó",
    "this message was deleted",
    "added you",
    "created this group",
    "salió del grupo",
}


def load_clients() -> List[str]:
    if not CLIENTS_FILE.exists():
        save_clients([BITACORA_DEFAULT_CLIENTE])
        return [BITACORA_DEFAULT_CLIENTE]
    try:
        data = json.loads(CLIENTS_FILE.read_text(encoding="utf-8") or "[]")
        if not isinstance(data, list):
            raise ValueError("invalid format")
        clean = [str(item).strip() for item in data if str(item).strip()]
    except Exception:
        clean = [BITACORA_DEFAULT_CLIENTE]
    if BITACORA_DEFAULT_CLIENTE not in clean:
        clean.append(BITACORA_DEFAULT_CLIENTE)
        save_clients(clean)
    return clean


def save_clients(items: List[str]) -> None:
    unique = []
    seen = set()
    for item in items:
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(item)
    CLIENTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    CLIENTS_FILE.write_text(json.dumps(unique, ensure_ascii=False, indent=2), encoding="utf-8")


class ProcessRequest(BaseModel):
    cliente: Optional[str] = None
    period: Optional[str] = None


class ClientCreate(BaseModel):
    nombre: str


class MetricResponse(BaseModel):
    total_registros: int
    clientes: Dict[str, int]
    estados: Dict[str, int]
    periodos: Dict[str, int]


def slugify(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", value or "sin_cliente")
    return cleaned.strip("_") or "sin_cliente"


def parse_date(value: str, fallback: datetime) -> datetime:
    if not value:
        return fallback
    for pattern in DATE_PATTERNS:
        try:
            return datetime.strptime(value.strip(), pattern)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value.strip())
    except Exception:
        return fallback


def normalize_field(key: str) -> Optional[str]:
    if not key:
        return None
    key_low = key.strip().lower()
    return FIELD_ALIASES.get(key_low)


def clean_input_line(value: str) -> str:
    if not value:
        return ""
    no_invisible = INVISIBLE_CHARS_RE.sub("", value)
    normalized_spaces = re.sub(r"\s+", " ", no_invisible)
    return normalized_spaces.strip()


def normalize_datetime(fecha_raw: str, hora_raw: str, ampm: str) -> str:
    candidates: List[Tuple[str, str]] = []

    # Intentos con AM/PM explícito
    if ampm:
        combined_time = f"{hora_raw} {ampm}".strip()
        for df in DATE_FORMATS:
            for tf in TIME_FORMATS_AMPM:
                candidates.append((f"{fecha_raw} {combined_time}".strip(), f"{df} {tf}"))

    # Intentos en formato 24h
    for df in DATE_FORMATS:
        for tf in TIME_FORMATS_24H:
            candidates.append((f"{fecha_raw} {hora_raw}".strip(), f"{df} {tf}"))

    for value, fmt in candidates:
        try:
            dt = datetime.strptime(value, fmt)
            year = dt.year
            if year < 2000:  # subir años de dos dígitos al rango 2000+
                year += 2000
                dt = dt.replace(year=year)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            continue

    logger.warning("No se pudo normalizar fecha/hora: '%s' '%s' '%s'", fecha_raw, hora_raw, ampm)
    return f"{fecha_raw} {hora_raw} {ampm}".strip()


def parse_whatsapp_line(line: str) -> Optional[Dict[str, str]]:
    cleaned_line = clean_input_line(line)
    if not cleaned_line:
        return None

    m = WHATSAPP_LINE_REGEX.match(cleaned_line)
    if not m:
        return None

    parts = m.groupdict()
    fecha_raw = parts.get("fecha", "").strip()
    hora_raw = parts.get("hora", "").strip()
    ampm = (parts.get("ampm") or "").strip().upper()
    autor = (parts.get("autor") or "").strip()
    mensaje = (parts.get("mensaje") or "").strip()

    lower_msg = mensaje.lower()
    if any(marker in lower_msg for marker in NOISE_KEYWORDS):
        return None

    fecha_iso = normalize_datetime(fecha_raw, hora_raw, ampm)
    try:
        fecha_dt = datetime.strptime(fecha_iso, "%Y-%m-%d %H:%M:%S")
    except Exception:
        fecha_dt = datetime.utcnow()

    return {
        "fecha": fecha_iso,
        "fecha_dt": fecha_dt,
        "cliente": "",
        "autor": autor,
        "mensaje": mensaje,
        "estado": "Pendiente",
    }


def parse_whatsapp_lines(lines: List[str]) -> List[Dict[str, str]]:
    records: List[Dict[str, str]] = []
    last_record: Optional[Dict[str, str]] = None
    for raw_line in lines:
        if not raw_line:
            continue
        parsed = parse_whatsapp_line(raw_line)
        if parsed:
            records.append(parsed)
            last_record = records[-1]
            continue

        # Soporta mensajes multilínea: si no hay fecha reconocible, concatena al anterior
        continuation = clean_input_line(raw_line)
        if continuation and last_record:
            last_record["mensaje"] = f"{last_record['mensaje']} {continuation}".strip()

    return records


def parse_structured_line(line: str) -> Optional[Dict[str, str]]:
    separators = [";", "|", ","]
    for sep in separators:
        if sep in line:
            parts = [p.strip() for p in line.split(sep)]
            if len(parts) >= 3:
                fecha_raw = parts[0]
                autor = parts[1]
                mensaje = parts[2]
                dummy_line = f"[{fecha_raw}, 00:00] {autor}: {mensaje}"
                return parse_whatsapp_line(dummy_line)
    return None


def parse_key_value_lines(lines: List[str]) -> Dict[str, str]:
    merged = " ".join([ln.strip() for ln in lines if ln.strip()])
    dummy_line = f"[01/01/2026, 00:00] anon: {merged}"
    parsed = parse_whatsapp_line(dummy_line)
    return parsed or {}


def sync_clients_from_entries(entries: List[Dict[str, str]]) -> List[str]:
    discovered = {BITACORA_DEFAULT_CLIENTE}
    for rec in entries:
        val = (rec.get("cliente") or BITACORA_DEFAULT_CLIENTE).strip()
        if val:
            discovered.add(val)
    current = set(load_clients())
    combined = sorted(current.union(discovered), key=lambda x: x.lower())
    save_clients(combined)
    return combined


def parse_txt_file(path: Path) -> Tuple[List[Dict[str, str]], Dict[str, int]]:
    stats = {"lines_read": 0, "lines_valid": 0}
    try:
        raw = path.read_text(encoding="utf-8", errors="ignore")
    except UnicodeDecodeError:
        raw = path.read_text(encoding="latin-1", errors="ignore")

    raw_lines = raw.splitlines()
    cleaned_non_empty = [clean_input_line(ln) for ln in raw_lines if clean_input_line(ln)]
    stats["lines_read"] = len(cleaned_non_empty)
    if not cleaned_non_empty:
        return [], stats

    whatsapp_records = parse_whatsapp_lines(raw_lines)
    if whatsapp_records:
        stats["lines_valid"] += len(whatsapp_records)
        return whatsapp_records, stats

    structured_records: List[Dict[str, str]] = []
    for line in cleaned_non_empty:
        parsed = parse_structured_line(line)
        if parsed:
            structured_records.append(parsed)
    if structured_records:
        stats["lines_valid"] += len(structured_records)
        return structured_records, stats

    parsed_kv = parse_key_value_lines(cleaned_non_empty)
    if parsed_kv:
        stats["lines_valid"] += 1
        return [parsed_kv], stats

    return [], stats


def collect_entries(source_dir: Path) -> Tuple[List[Dict[str, str]], Dict[str, int]]:
    entries: List[Dict[str, str]] = []
    agg_stats = {"lines_read": 0, "lines_valid": 0, "records": 0}
    for file_path in glob.glob(str(source_dir / "*.txt")):
        path_obj = Path(file_path)
        parsed_records, stats = parse_txt_file(path_obj)
        agg_stats["lines_read"] += stats.get("lines_read", 0)
        agg_stats["lines_valid"] += stats.get("lines_valid", 0)
        for rec in parsed_records:
            if not rec.get("cliente"):
                rec["cliente"] = BITACORA_DEFAULT_CLIENTE
            entries.append(rec)
            agg_stats["records"] += 1
    return entries, agg_stats


def period_from_date(date_obj: datetime) -> str:
    return date_obj.strftime("%Y-%m")




def build_metrics(entries: List[Dict[str, str]]) -> MetricResponse:
    clientes: Dict[str, int] = defaultdict(int)
    estados: Dict[str, int] = defaultdict(int)
    periodos: Dict[str, int] = defaultdict(int)
    for e in entries:
        clientes[e.get("cliente", "SIN_CLIENTE")]+=1
        estados[e.get("estado", "") or "Sin estado"]+=1
        periodos[period_from_date(e.get("fecha_dt", datetime.utcnow()))]+=1
    return MetricResponse(
        total_registros=len(entries),
        clientes=dict(clientes),
        estados=dict(estados),
        periodos=dict(periodos),
    )


def write_excel(rows: List[Dict[str, str]], cliente: str, year: int) -> Path:
    output_dir = BITACORA_OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_cliente = slugify(cliente).lower()
    file_name = f"bitacora_{safe_cliente}_{year}.xlsx"
    full_path = output_dir / file_name

    if full_path.exists():
        full_path.unlink()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bitacora"

    headers = [
        "Fecha",
        "Cliente",
        "Reportado por",
        "Descripción",
        "Estado",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="345995", end_color="345995", fill_type="solid")
    center = Alignment(vertical="center")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    zebra_fill = PatternFill(start_color="F8FBFF", end_color="F8FBFF", fill_type="solid")

    for idx, r in enumerate(rows, start=2):
        ws.append([
            r.get("fecha"),
            r.get("cliente"),
            r.get("autor"),
            r.get("mensaje"),
            r.get("estado"),
        ])
        if idx % 2 == 0:
            for col in range(1, len(headers)+1):
                ws.cell(row=idx, column=col).fill = zebra_fill

    for col in range(1, len(headers)+1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

    wb.save(full_path)
    return full_path


app = FastAPI(title="Bitacoras Mantenimiento", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"]
)


@app.get("/health")
def health():
    return {"status": "ok", "source": str(BITACORA_SOURCE_DIR), "output": str(BITACORA_OUTPUT_DIR)}


@app.post("/bitacora/procesar")
def procesar(
    req: ProcessRequest | None = Body(default=None)
):
    if not BITACORA_SOURCE_DIR.exists():
        return {
            "success": True,
            "status": "ok",
            "message": "No hay nuevos registros para procesar",
            "generated": []
        }

    messages, stats = collect_entries(BITACORA_SOURCE_DIR)
    if not messages:
        return {
            "success": True,
            "status": "ok",
            "message": "No hay nuevos registros para procesar",
            "generated": [],
            "lines_read": stats.get("lines_read", 0),
            "lines_valid": stats.get("lines_valid", 0),
            "new_records": 0,
        }

    target_cliente = ((req.cliente if req else None) or BITACORA_DEFAULT_CLIENTE).strip() or BITACORA_DEFAULT_CLIENTE
    target_period = (req.period if req else None) or None

    try:
        target_year = int(target_period.split("-")[0]) if target_period else datetime.utcnow().year
    except Exception:
        target_year = datetime.utcnow().year

    for e in messages:
        if not e.get("cliente") or e["cliente"] == BITACORA_DEFAULT_CLIENTE:
            e["cliente"] = target_cliente

    filtered = []
    for e in messages:
        if e.get("cliente", "").lower() != target_cliente.lower():
            continue
        period_val = period_from_date(e.get("fecha_dt", datetime.utcnow()))
        if target_period and period_val != target_period:
            continue
        if not target_period and e.get("fecha_dt", datetime.utcnow()).year != target_year:
            continue
        filtered.append(e)

    if not filtered:
        return {
            "success": True,
            "status": "ok",
            "message": "No hay nuevos registros para procesar",
            "generated": [],
            "lines_read": stats.get("lines_read", 0),
            "lines_valid": stats.get("lines_valid", 0),
            "new_records": 0,
        }

    sync_clients_from_entries(filtered)

    excel_path = write_excel(filtered, target_cliente, target_year)
    metrics = build_metrics(filtered)
    new_records = len(filtered)

    discarded = stats.get("lines_read", 0) - stats.get("lines_valid", 0)
    logger.info(
        "Bitacora procesada | Leidas: %s | Parseadas: %s | Descartadas: %s | Nuevas agregadas: %s",
        stats.get("lines_read", 0),
        stats.get("lines_valid", 0),
        discarded,
        new_records,
    )

    return {
        "success": True,
        "generated": [{
            "cliente": target_cliente,
            "periodo": target_period or str(target_year),
            "rows": len(filtered),
            "path": str(excel_path),
            "filename": excel_path.name,
        }],
        "metrics": metrics.dict(),
        "processed_at": datetime.utcnow().isoformat(),
        "request_id": str(uuid.uuid4()),
        "message": "Bitácora actualizada",
        "lines_read": stats.get("lines_read", 0),
        "lines_valid": stats.get("lines_valid", 0),
        "new_records": new_records,
        "status": "ok",
    }


@app.get("/bitacora/metricas", response_model=MetricResponse)
def metricas():
    messages, _ = collect_entries(BITACORA_SOURCE_DIR)
    for msg in messages:
        if not msg.get("cliente"):
            msg["cliente"] = BITACORA_DEFAULT_CLIENTE
    sync_clients_from_entries(messages)
    return build_metrics(messages)


@app.get("/bitacora/clientes")
def listar_clientes():
    clientes = load_clients()
    return {"clientes": clientes, "default": BITACORA_DEFAULT_CLIENTE}


@app.post("/bitacora/clientes")
def crear_cliente(body: ClientCreate):
    nombre = (body.nombre or "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre de cliente es obligatorio")

    clientes = load_clients()
    nombres_lc = [c.lower() for c in clientes]
    if nombre.lower() in nombres_lc:
        return {"success": False, "message": "El cliente ya existe", "clientes": clientes, "created": False}

    clientes.append(nombre)
    save_clients(clientes)
    return {"success": True, "message": "Cliente creado correctamente", "clientes": sorted(clientes, key=str.lower), "created": True}


@app.get("/bitacora/excel")
def descargar(cliente: Optional[str] = None, period: Optional[str] = None):
    target_cliente = (cliente or BITACORA_DEFAULT_CLIENTE).strip() or BITACORA_DEFAULT_CLIENTE
    safe_cliente = slugify(target_cliente).lower()
    try:
        target_year = int(period.split("-")[0]) if period else datetime.utcnow().year
    except Exception:
        target_year = datetime.utcnow().year

    chosen_path = BITACORA_OUTPUT_DIR / f"bitacora_{safe_cliente}_{target_year}.xlsx"

    if not chosen_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado. Procese la bitácora primero.")

    return FileResponse(
        path=chosen_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=chosen_path.name,
    )


@app.exception_handler(Exception)
async def unhandled_exception_handler(request, exc):
    return JSONResponse(status_code=500, content={"error": "internal_error", "detail": str(exc)[:200]})


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "main:app",
        host="127.0.0.1",
        port=8000,
        reload=False
    )
